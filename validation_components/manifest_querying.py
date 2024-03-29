import xlrd
import openpyxl
import requests
from datetime import datetime
import time


class ManifestEntry:
    '''
    A class construct for a single entry from the manifest submitted and query the entries when necessary
    '''
    def __init__(self, sample_id: str, common_name, taxon_id):
        self.sample_id = sample_id
        self.common_name = common_name
        self.taxon_id = taxon_id
        self.query_id = common_name + str(taxon_id)

    @staticmethod
    def report_error(error_code, common_name_statement, taxon_id_statement):
        if error_code == 1:
            error = (f": Taxon ID and common name don't match. " + common_name_statement
                     + taxon_id_statement)
        elif error_code == 2:
            error = f": {taxon_id_statement}"
        else:
            error = (f": " + common_name_statement + taxon_id_statement)
        return error

    def common_name_definition(self, ncbi_result):
        if self.common_name == '__null__':
            statement = 'No common name specified. '
        elif ncbi_result == '__null__':
            statement = f"The given common name '{self.common_name}' does not exist in the NCBI database. "
        else:
            statement = f"The taxon ID for given name '{self.common_name}' is {ncbi_result}. "
        return statement

    def taxon_id_definition(self, ncbi_result=None):
        if self.taxon_id == '__null__':
            statement = 'No taxon ID specified.'
        elif ncbi_result == '__null__':
            statement = f'The given taxon ID {self.taxon_id} does not exist in the NCBI database.'
        else:
            statement = f"The official name for the given taxon ID {self.taxon_id} is '{ncbi_result}'."
        return statement

class NcbiQuery:

    def __init__(self):
        self.timestamp = None

    @staticmethod
    def get_now():
        return datetime.now()

    def generate_new_timestamp(self):
        if self.timestamp != None:
            time_since_last_query = NcbiQuery.get_now() - self.timestamp

            required_microsecond_delay = 335000
            if time_since_last_query.days == 0 and time_since_last_query.seconds == 0 and time_since_last_query.microseconds < required_microsecond_delay:
                time.sleep((required_microsecond_delay - time_since_last_query.microseconds) / 1000000)

        self.timestamp = NcbiQuery.get_now()


    def query_ncbi_for_taxon_id(self, manifest_entry: ManifestEntry):
        self.generate_new_timestamp()
        url = self.build_url(manifest_entry, esearch=True)
        tax_id_json = self.ncbi_search(url)
        if 'esearchresult' in tax_id_json and 'idlist' in tax_id_json['esearchresult'] and len(
                tax_id_json['esearchresult']['idlist']) == 1:
            return str(tax_id_json['esearchresult']['idlist'][0])
        else:
            return '__null__'

    def query_ncbi_for_common_name(self, manifest_entry: ManifestEntry):
        self.generate_new_timestamp()
        url = self.build_url(manifest_entry, esearch=False)
        common_name_json = self.ncbi_search(url)
        if 'result' in common_name_json and manifest_entry.taxon_id in common_name_json['result'] and 'scientificname' in \
                common_name_json['result'][manifest_entry.taxon_id]:
            ncbi_common_name = common_name_json['result'][manifest_entry.taxon_id]['scientificname']
            ncbi_ranking = common_name_json['result'][manifest_entry.taxon_id]['rank']
        else:
            ncbi_common_name = '__null__'
            ncbi_ranking = None
        return ncbi_common_name, ncbi_ranking

    def build_url(self, manifest_entry, esearch: bool):
        base_url = 'https://eutils.ncbi.nlm.nih.gov/entrez/eutils/'
        mid_url = '.fcgi?db=taxonomy&'
        end_url = '&retmode=json'
        if esearch:
            url = base_url + 'esearch' + mid_url + 'field=All%20Names&term=' + manifest_entry.common_name + end_url
        else:
            url = base_url + 'esummary' + mid_url + 'id=' + str(manifest_entry.taxon_id) + end_url
        return url


    def ncbi_search(self, url):
        search_session = requests.Session()
        data = search_session.get(url)
        if not data:
            raise ConnectionError('Could not connect to NCBI database - Aborting')
        return data.json()


class SpreadsheetLoader:

    def __init__(self, file):
        self._file = file
        try:
            self._workbook = openpyxl.load_workbook(self._file)
            self._sheet = self._workbook.worksheets[0]
            self._format = 'xlsx'
        except:
            self._workbook = xlrd.open_workbook(self._file)
            self._sheet = self._workbook.sheet_by_index(0)
            self._format = 'xls'

    def load(self):
        entries = []
        data_row = 0
        header_row = 0
        for row in range(self._sheet.nrows):
            if self._sheet.cell_value(row, 0) == 'SANGER PLATE ID':
                data_row = row + 1
                header_row = row
                break

        for column in range(self._sheet.ncols):
            header_cell = self._sheet.cell_value(header_row, column)
            if header_cell == 'SUPPLIER SAMPLE NAME':
                sample_id_column = column
            elif header_cell == 'TAXON ID':
                taxon_id_column = column
            elif header_cell == 'COMMON NAME':
                common_name_column = column


        for row in range(data_row, self._sheet.nrows):
            common_name = self.__extract_cell_value(row, common_name_column)
            taxon_id = self.__extract_cell_value(row, taxon_id_column)
            sample_id = self.__extract_cell_value(row, sample_id_column)
            if sample_id != '__null__':
                entry = ManifestEntry(sample_id, common_name, taxon_id)
                entries.append(entry)
        return entries


    def __extract_cell_value(self, row, column):
        if self._sheet.cell_type(row, column) != xlrd.XL_CELL_NUMBER:
            new_data = self._sheet.cell_value(row, column).strip()
            return '__null__' if new_data == '' else new_data.replace('\xa0',' ')
        new_data = str(int(self._sheet.cell_value(row, column)))
        return '__null__' if new_data == '' else new_data

    def load_xlsx(self):
        entries = []
        data_row = 0
        header_row = 0
        for row in range(self._sheet.max_row):
            if self._sheet.cell(row=row + 1, column=1).value == 'SANGER PLATE ID':
                data_row = row + 2
                header_row = row + 1
                break


        for col in range(self._sheet.max_column):
            header_cell = self._sheet.cell(row=header_row, column=col + 1).value
            if header_cell == 'SUPPLIER SAMPLE NAME':
                sample_id_column = col + 1
            elif header_cell == 'TAXON ID':
                taxon_id_column = col + 1
            elif header_cell == 'COMMON NAME':
                common_name_column = col + 1
        reads = []

        for row in range(data_row, self._sheet.max_row+1):
            common_name = self.__extract_cell_value_xlsx(row, common_name_column)
            taxon_id = self.__extract_cell_value_xlsx(row, taxon_id_column)
            sample_id = self.__extract_cell_value_xlsx(row, sample_id_column)
            if sample_id != '__null__':
                entry = ManifestEntry(sample_id, common_name, taxon_id)
                entries.append(entry)
        return entries


    def __extract_cell_value_xlsx(self, row, column):
        if isinstance(self._sheet.cell(row=row, column=column).value, str):
            new_data = self._sheet.cell(row=row, column=column).value.strip()
            return '__null__' if new_data == '' else new_data.replace('\xa0',' ')
        new_data = self._sheet.cell(row=row, column=column).value
        return '__null__' if new_data == None else str(int(new_data))
